# 京东商品下单

from types import DynamicClassAttribute
import unittest
import openpyxl
from unittest.main import main
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By


class TestJD(unittest.TestCase):

    def setUp(self) -> None:
         # 全局配置
        global driver

        # 加载浏览器
        driver = webdriver.Chrome()

        # 窗口最大化
        driver.maximize_window()

        # 加载地址
        driver.get("http://t-mintegral437.helitong.cn/#/Login")
        # 等待时间
        sleep(5)


        """登录"""
        #输入账号
        driver.find_element(By.XPATH, "//input[@name='mobile']/following::input[2]").send_keys("13017167459")
        sleep(1)
        # 输入密码
        driver.find_element(By.NAME, "password").send_keys("t123456")
        sleep(1)
        # 点击登录
        driver.find_element(By.XPATH, "//button/following::div[10]").click()
        sleep(2)

    def test_jdGoods(self):
       
        while True:

            """购物车"""
            # 点击购物车
            driver.find_element(By.XPATH, "//div[@class='van-tabbar-item__text']/following::span[2]").click()
            sleep(1)
            """
            点击第一个勾选框
            判断是否存在元素
            默认存在
            """
            element_existence = True
            try:
                # 定位元素尝试点击
                driver.find_element(By.XPATH, "//div[@role='checkbox']/following::div[4]").click()
                sleep(1)
                # 没有定位到元素则抛出异常
            except:
                element_existence = False

            # 判断元素是否存在，不存在则跳出循环，存在则继续循环
            if element_existence == False:
                break
            else:                 
                # 结算
                driver.find_element(By.XPATH, "//div[span=' 全选 ']/following::button").click()
                sleep(2)

                """结算页面"""
                # 选择使用积分
                driver.find_element(By.XPATH, "//div[span='去选择']").click()
                sleep(1)

                # 确认使用积分
                driver.find_element(By.XPATH, "//span[@class='allintegralcolor']/following::button").click()
                sleep(1)

                # 提交订单
                driver.find_element(By.XPATH, "//span[@class='allprice']/following::button").click()
                sleep(5)

                # 输入密码
                driver.find_element(By.XPATH, "//input[@class='password-code-input']").send_keys("201518")
                sleep(1)

                # 确认
                driver.find_element(By.XPATH, "//div[@class='buttons-bar']/button[2]").click()
                sleep(3)

                element_click = True
                try:
                    # 查看订单
                    driver.find_element(By.XPATH, "//div[@class='success']//button[2]").click()
                    sleep(1)

                    # 点击商品详情
                    driver.find_element(By.XPATH, "//div[span='待发货']/following::span[5]").click()
                    sleep(1)

                    # 获取订单详情
                    order_sn = driver.find_element(By.XPATH, "//div[@class='van-cell__value']/following::div[37]").text
                    goods_name = driver.find_element(By.XPATH, "//div[@class='van-card__content']//div/div").text
                    print(f"下单成功：\t订单编号：{order_sn}\t商品名称：{goods_name}")
                    file_book = open("C:/Users/Administrator/Desktop/1.txt", mode="a", encoding="utf-8")
                    file_book.write(f"{order_sn}\t")
                    file_book.write(f"{goods_name}\n")
                    file_book.close()

                    def write_excel():
                        work_book = openpyxl.Workbook()
                        sheet = work_book.create_sheet("new")
                        data = open("C:/Users/Administrator/Desktop/1.txt", "r", encoding="utf-8")
                        datas = data.readlines()
                        for i,row in enumerate(datas):
                            d = row.split()
                            for j in range(len(d)):
                                sheet.cell(i+1, j+1, d[i])
                                work_book.save("C:/Users/Administrator/Desktop/output.xlsx")
                    write_excel()
                   

                    # 返回上一步
                    driver.find_element(By.XPATH, " //div[@class='van-cell']").click()
                    sleep(1)

                    # 退回上一步
                    driver.back()
                               
                    """回到首页"""                
                    driver.find_element(By.XPATH, "//div[@class='success']//button").click()
                    sleep(2)
                except:
                    element_click = False
                if element_click == False:
                    # 点击取消
                    driver.find_element(By.XPATH, " //div[@class='buttons-bar']//button").click()
                    sleep(1)
                    # 点击返回
                    driver.find_element(By.XPATH, " //div[@class='van-cell']").click()
                    sleep(1)

                    """获取商品名称"""
                    text = driver.find_element(By.XPATH, " //div[@class='van-card__content']//div/div").text
                    print(f"下单失败：\t商品名称：{text}")

                    file_book = open("C:/Users/Administrator/Desktop/1.txt", mode="a", encoding="utf-8")
                    file_book.write(f"{text}\n")
                    file_book.close()

                    def write_excel():
                        work_book = openpyxl.Workbook()
                        sheet = work_book.create_sheet("new")
                        data = open("C:/Users/Administrator/Desktop/1.txt", "r", encoding="utf-8")
                        datas = data.readlines()
                        for i,row in enumerate(datas):
                            d = row.split()
                            for j in range(len(d)):
                                sheet.cell(i+1, j+1, d[i])
                                work_book.save("C:/Users/Administrator/Desktop/output.xlsx")
                    write_excel()

                    # 点击编辑
                    driver.find_element(By.XPATH, " //div[span='编辑']").click()
                    sleep(1)
                    # 选中第一个商品
                    driver.find_element(By.XPATH, " //div[@class='van-swipe-cell']//div/div").click()
                    sleep(1)
                    # 点击删除
                    driver.find_element(By.XPATH, " //div[@class='delete-bar']//button").click()
                    sleep(1)
                    # 点击确认
                    driver.find_element(By.XPATH, " //div[@class='van-dialog']//button[2]").click()
                    sleep(2)
                else:
                    # 跳出循环
                    continue


    # 用例执行完关闭页面
    def tearDown(self) -> None:
        driver.close()


if __name__ == "__main__":
    unittest.main()